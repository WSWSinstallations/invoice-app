from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


HEADERS = [
    "supplier",
    "invoice number",
    "date",
    "project",
    "item",
    "qty",
    "price",
    "total",
    "category",
]


def generate_invoice_excel(invoice, output_path: str | Path) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Invoice line items"
    sheet.append(HEADERS)

    for line in invoice.line_items:
        sheet.append(
            [
                invoice.supplier,
                invoice.invoice_number,
                invoice.invoice_date,
                invoice.project,
                line.item,
                line.qty,
                line.price,
                line.total,
                line.category,
            ]
        )

    header_fill = PatternFill("solid", fgColor="E6F4F1")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    for row in sheet.iter_rows(min_row=2, min_col=6, max_col=8):
        for cell in row:
            cell.number_format = "#,##0.00"

    for column_cells in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = min(max(max_length + 2, 12), 28)

    Path(output_path).parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)
